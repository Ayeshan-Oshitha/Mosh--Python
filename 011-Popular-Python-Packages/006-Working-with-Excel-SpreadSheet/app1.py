# ---------------
# This script demonstrates a few basic methods
# ---------------

import openpyxl

# Load an existing Excel workbook
wb = openpyxl.load_workbook("Files6/transactions.xlsx")

print(wb.sheetnames)

# Select the worksheet named "Sheet1"
sheet = wb["Sheet1"]

# Create a new sheet named "Sheet2" and insert it at the first position (index 0)
wb.create_sheet("Sheet2", 0)

# Access a cell using A1 notation
cell = sheet["a1"]
print(cell.row)
print(cell.column)
print(cell.coordinate)

# Access a cell using row and column indices
cell = sheet.cell(row=1, column=1)

cell = sheet["a1"]

# Loop through all rows and columns and print each cell's value
for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        cell = sheet.cell(row, column)
        print(cell.value)


# Append a new row of data at the end of the worksheet
sheet.append([1, 2, 3])

# Save the modified workbook to a new file
wb.save("Files6/transactions2.xlsx")
